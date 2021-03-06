import os
import re
import sys
import datetime
import openpyxl
from openpyxl.chart import BarChart, Reference, Series

def close_book(wb, filepath):
    # 종료
    wb.save(filepath)
    wb.close()
########################################################
### 업무지원비 서류 작업
########################################################

# 거래내역 위치 좌상단 B2부터 우상단 I2
# 승인일시(B2) 승인번호(C2) 가맹점번호(D2) 
# 가맹점명(E2) 승인금액(F2) 매출종류(G2) 
# 접수월일(H2) 승인취소(I2)
fields1_range = range(2,10)
# 카드번호(B2) 승인번호(C2) 처리일(D2) 매출일(E2) 
# 종류(F2) 가맹점(G2) 현지금액(H2) 결제금액($)(I2) 
# 이용금액(W)(J2) 국가(K2) 도시(L2)
fields2_range = range(2,13)

# 날짜 및 시간 지정 문자열	의미
# %Y	앞의 빈자리를 0으로 채우는 4자리 연도 숫자
# %m	앞의 빈자리를 0으로 채우는 2자리 월 숫자
# %d	앞의 빈자리를 0으로 채우는 2자리 일 숫자
# %H	앞의 빈자리를 0으로 채우는 24시간 형식 2자리 시간 숫자
# %M	앞의 빈자리를 0으로 채우는 2자리 분 숫자
# %S	앞의 빈자리를 0으로 채우는 2자리 초 숫자
# %A	영어로 된 요일 문자열
# %B	영어로 된 월 문자열

class Record:
    '''
        brief: It is a Data Class of Money book.
    '''
    def __init__(self, date, money, where):
        # 자주 사용되는 장소
        purpose_dict = {
            "주유소":	"주유비",
            "미니스톱":	"간식 구매",
            "GS25":	"간식 구매",
            "유니스토아":	"물품 구매",
            "이마트":	"연구소 간식 구매",
            "휴게소":	"식사 구매",
            "에스엠하이플러스":	"교통비",
            "씨유":	"간식 구매",
            "씨앤에스자산관리":	"주차비",
            "티머니택시":	"교통비",
            "호반베르디움아브뉴프랑판교지점":	"주차비",
            "쿠팡":	"물품 구매",
            "Amazon Prime" : "배송비",
            "ARETHUSA FARM DAIRY" : "간식 구매",
            "BURGER KING" : "식사비",
            "CHIPOTLE" : "영업처 직원들과 회의",
            "COSTCO GAS" : "주유비",
            "COSTCO WHSE" : "간식 구매",
            "IKEA" : "소모품 구매",
            "MTA*MNR STATION" : "교통비",
            "OLIVE GARDEN" : "영업처 직원들과 회의",
            "RIVERDALE DINER" : "영업처 직원들과 회의",
            "SHAKE SHACK" : "식사비",
            "SHELL OIL" : "주유비",
            "STARBUCKS" : "커피 구매",
            "SUBWAY" : "식사 구매",
            "WHOLEFDS MIL" : "간식 구매"
        }
        self.date = date
        self.money = money
        print(where)
        # 사용처 길이가 짧을 경우 그대로 사용
        if len(where) < 10:
            self.where = where
        # 사용처 길이가 길 경우 줄이기
        else:
            where = re.sub(r'\(.+\)',"",where)
            self.where = where.replace("주식회사","")

        # 사용 목적 처리 : 1. 즐겨찾기, 2. 직접입력
        has_purpose = False
        for key in purpose_dict.keys():
            if key in self.where:
                has_purpose = has_purpose or True
                self.purpose = purpose_dict.get(key)
        if not has_purpose:
            self.purpose = "** 직접입력하세요. **"
        

    def __lt__(self, other):
             return self.date < other.date

###
def handle_abroad(worksheet):
    '''
        brief: handle abroad money records
    '''
    records = []
    for idx in range(3, worksheet.max_row+1):
        if worksheet.cell(idx,5).value is not  None:
            dt = worksheet.cell(idx,5).value
            record = Record(dt, worksheet.cell(idx,10).value, worksheet.cell(idx,7).value)
            records.append(record)
        else:
            print("국외 결제 레코드 총 갯수 : ", idx-3) # 첫째줄, 제목줄, 마지막 공백 줄
            break
    records.sort()
    return records

def handle_domestic(worksheet):
    '''
        brief: handle domestic money records
    '''
    records = []
    for idx in range(3, worksheet.max_row+1):
        if worksheet.cell(idx,2).value is not  None:
            # dt = datetime.datetime.strptime(str(worksheet.cell(idx,2).value), "%Y-%m-%d %H:%M:%S")
            dt = worksheet.cell(idx,2).value
            record = Record(dt, worksheet.cell(idx,6).value, worksheet.cell(idx,5).value)
            records.append(record)
        else:
            print("국내 결제 레코드 총 갯수 : ", idx-3) # 첫째줄, 제목줄, 마지막 공백 줄
            break
    records.sort()
    return records

###
def write_records(to_records, from_records, record_idx):
    '''
        brief: write records to a excel-worksheet
    '''
    for record in from_records:
        to_records.cell(record_idx,2).value = record.date.strftime('%m.%d')
        to_records.cell(record_idx,3).value = record.money
        to_records.cell(record_idx,4).value = record.where
        to_records.cell(record_idx,5).value = record.purpose
        record_idx += 1
    return record_idx



### List & Filter the workbooks
def excel_list(dirname):
    '''
        brief: lists up of a directory
    '''
    excel_list = []
    filename_type = r'^.+\.xlsx'
    for filename in os.listdir(dirname):
        if re.search(filename_type, filename):
            excel_list.append(filename)
    return excel_list

def filter_worksheet(dirname, filelist):
    '''
        brief: filter a list with the cost worksheet condition
    '''
    checked_list = []
    for filename in filelist:
        filepath = os.path.join(dirname, filename)
        wb = openpyxl.load_workbook(filepath)
        if all(item in wb.sheetnames for item in ['국내','국외']):
            if not '종합' in wb.sheetnames:
                checked_list.append(filename)
    return checked_list

def find_checked_excel_list(dirname):
    '''
        brief: lists up & filter excel workbooks of a directory
    '''
    excel_filelist = excel_list(dirname)
    checked_excel_list = filter_worksheet(dirname, excel_filelist)
    return checked_excel_list

### Sum up to new worksheet
def process_excel_costs(filepath):
    '''
        brief: process(sum up) worksheets; '국내', '국외' to '종합'
    '''
    if os.path.exists(filepath):
        print("Start Pasing...")
        wb = openpyxl.load_workbook(filepath)
        domestic = wb['국내']
        abroad = wb['국외']
        domestic_records = handle_domestic(domestic)
        abroad_records = handle_abroad(abroad)

        # 종합 정리
        total = wb.create_sheet('sheet3')
        total.title = '종합'
        record_idx = 3
        record_idx = write_records(total, domestic_records, record_idx)
        record_idx += 1
        record_idx = write_records(total, abroad_records, record_idx)

        return wb, record_idx

### Chart
def draw_barchart(worksheet, last_idx):
    values = Reference(worksheet, min_col=2, min_row=2, max_col=3, max_row=last_idx)
    chart = BarChart()
    chart.add_data(values)
    worksheet.add_chart(chart, "H2")


##########################################################
# 메인
##########################################################

download_dir = "/Users/taeyoonlee/Downloads/"
checked_excel_list = find_checked_excel_list(download_dir)

for excel in checked_excel_list:
    filepath = os.path.join(download_dir, excel)
    each_workbook, each_last_record_idx = process_excel_costs(filepath)
    draw_barchart(each_workbook['종합'], each_last_record_idx)
    close_book(each_workbook, filepath)


